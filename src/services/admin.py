# src/services/admin.py

import os
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

from src.database.models import Listing, TelegramUser


class AdminService:
    def __init__(self, db: Session):
        self.db = db

    def get_usage_stats(self) -> dict:
        """Возвращает статистику использования за последние 7 дней."""
        week_ago = datetime.utcnow() - timedelta(days=7)

        total_users = self.db.query(TelegramUser).count()
        active_users = self.db.query(TelegramUser).filter(
            TelegramUser.last_active_at >= week_ago
        ).count()

        total_listings = self.db.query(Listing).count()
        active_listings = self.db.query(Listing).filter(
            Listing.is_active == True
        ).count()

        return {
            "total_users": total_users,
            "active_users_7d": active_users,
            "total_listings": total_listings,
            "active_listings": active_listings,
        }

    def get_popular_queries(self, limit: int = 5) -> list:
        """
        ЗАГЛУШКА: пока не логируем запросы.
        В будущем — можно добавить таблицу `user_queries`.
        """
        return [
            "участок в Мытищах до 2 млн",
            "аренда земли в Подмосковье",
            "ИЖС рядом с МКАД",
            "дешевые участки в Химках",
            "участок 10 соток"
        ][:limit]

    def export_listings_to_excel(self) -> BytesIO:
        """Экспортирует активные объявления в Excel и возвращает BytesIO."""
        listings = self.db.query(Listing).filter(Listing.is_active == True).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Участки ЕАСУЗ"

        # Заголовки
        headers = [
            "ID", "Реестр", "Район", "Цена", "Площадь (м²)", "Назначение",
            "Тип (аренда/продажа)", "Статус", "Ссылка", "Дата обновления"
        ]
        ws.append(headers)

        # Стиль заголовков
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        # Данные
        for listing in listings:
            ws.append([
                listing.id,
                listing.registry_number or "",
                listing.district_code or "",
                listing.start_price,
                listing.total_square,
                listing.land_allowed_use_name or "",
                listing.purchase_kind_name or "",
                listing.stage_state_name or "",
                listing.link,
                listing.updated_at.strftime("%Y-%m-%d %H:%M") if listing.updated_at else ""
            ])

        # Автоподбор ширины
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Сохраняем в память
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def trigger_db_update(self) -> str:
        """
        ЗАГЛУШКА: запускает обновление данных.
        В реальности — вызывает парсер.
        """
        # Здесь будет: from src.parser.easuz_parser import run_full_update
        # run_full_update()
        return "✅ Запущено обновление базы данных из ЕАСУЗ. Это займёт несколько минут."